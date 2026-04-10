"""
Export service for generating comparison reports.
"""
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import csv


class ExportService:
    """Generate and export comparison reports."""
    
    @staticmethod
    def export_to_excel(comparison_result, file_info1, file_info2):
        """
        Export comparison results to Excel workbook.
        
        Returns:
            io.BytesIO: Excel file buffer
        """
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Create worksheets
        ExportService._create_summary_sheet(workbook, comparison_result, file_info1, file_info2)
        ExportService._create_matched_rows_sheet(workbook, comparison_result)
        ExportService._create_unmatched_sheet(workbook, comparison_result)
        ExportService._create_differences_sheet(workbook, comparison_result)
        
        # Write to buffer
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def _create_summary_sheet(workbook, comparison_result, file_info1, file_info2):
        """Create summary worksheet."""
        ws = workbook.create_sheet('Summary', 0)
        
        # Header styling
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        # Title
        ws['A1'] = 'UAT Data Comparison Report'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        # File information
        ws['A3'] = 'File 1 (Developer Upload)'
        ws['A3'].font = Font(bold=True)
        ws['A4'] = f"Name: {file_info1.get('original_filename', 'N/A')}"
        ws['A5'] = f"Size: {file_info1.get('file_size', 0) / 1024:.2f} KB"
        ws['A6'] = f"Uploaded: {file_info1.get('uploaded_at', 'N/A')}"
        
        ws['A8'] = 'File 2 (Client Upload)'
        ws['A8'].font = Font(bold=True)
        ws['A9'] = f"Name: {file_info2.get('original_filename', 'N/A')}"
        ws['A10'] = f"Size: {file_info2.get('file_size', 0) / 1024:.2f} KB"
        ws['A11'] = f"Uploaded: {file_info2.get('uploaded_at', 'N/A')}"
        
        # Statistics
        stats = comparison_result.get('statistics', {})
        ws['A13'] = 'Comparison Statistics'
        ws['A13'].font = Font(bold=True)
        
        ws['A14'] = 'Metric'
        ws['B14'] = 'Value'
        for row, (key, val) in enumerate(stats.items(), start=15):
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = val
        
        # Quality score
        quality_score = comparison_result.get('quality_score', 0)
        ws['A24'] = 'Overall Quality Score'
        ws['A24'].font = Font(bold=True)
        ws['B24'] = f"{quality_score}%"
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    @staticmethod
    def _create_matched_rows_sheet(workbook, comparison_result):
        """Create matched rows worksheet."""
        ws = workbook.create_sheet('Matched Rows')
        
        matched_rows = comparison_result.get('rows', {}).get('matched_rows', [])
        
        # Headers
        ws['A1'] = 'File 1 Row'
        ws['B1'] = 'File 2 Row'
        ws['C1'] = 'Similarity'
        ws['D1'] = 'Differences'
        
        # Data
        for idx, row in enumerate(matched_rows, start=2):
            ws[f'A{idx}'] = row.get('file1_row')
            ws[f'B{idx}'] = row.get('file2_row')
            ws[f'C{idx}'] = row.get('similarity')
            ws[f'D{idx}'] = len(row.get('differences', []))
        
        # Set column widths
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 15
    
    @staticmethod
    def _create_unmatched_sheet(workbook, comparison_result):
        """Create unmatched rows worksheet."""
        ws = workbook.create_sheet('Unmatched')
        
        row_diff = comparison_result.get('rows', {})
        unmatched_f1 = row_diff.get('unmatched_in_file1', [])
        unmatched_f2 = row_diff.get('unmatched_in_file2', [])
        
        # File 1 unmatched
        ws['A1'] = 'File 1 - Unmatched Rows'
        ws['A1'].font = Font(bold=True)
        
        for idx, row in enumerate(unmatched_f1, start=2):
            ws[f'A{idx}'] = f"Row {row.get('row_index')}"
        
        # File 2 unmatched
        col_offset = 5
        ws[f'{chr(65 + col_offset)}1'] = 'File 2 - Unmatched Rows'
        ws[f'{chr(65 + col_offset)}1'].font = Font(bold=True)
        
        for idx, row in enumerate(unmatched_f2, start=2):
            ws[f'{chr(65 + col_offset)}{idx}'] = f"Row {row.get('row_index')}"
        
        ws.column_dimensions['A'].width = 30
    
    @staticmethod
    def _create_differences_sheet(workbook, comparison_result):
        """Create detailed differences worksheet."""
        ws = workbook.create_sheet('Differences')
        
        matched_rows = comparison_result.get('rows', {}).get('matched_rows', [])
        
        # Headers
        ws['A1'] = 'File 1 Row'
        ws['B1'] = 'File 2 Row'
        ws['C1'] = 'Column'
        ws['D1'] = 'File 1 Value'
        ws['E1'] = 'File 2 Value'
        
        # Data
        row_idx = 2
        for row in matched_rows:
            for diff in row.get('differences', []):
                ws[f'A{row_idx}'] = row.get('file1_row')
                ws[f'B{row_idx}'] = row.get('file2_row')
                ws[f'C{row_idx}'] = diff.get('column')
                ws[f'D{row_idx}'] = diff.get('file1_value')
                ws[f'E{row_idx}'] = diff.get('file2_value')
                row_idx += 1
        
        # Set column widths
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 25
    
    @staticmethod
    def export_to_csv(comparison_result, file_info1, file_info2):
        """
        Export summary to CSV format.
        
        Returns:
            str: CSV content
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(['UAT Data Comparison Report'])
        writer.writerow([])
        
        # File info
        writer.writerow(['File Information'])
        writer.writerow(['File 1 (Developer):', file_info1.get('original_filename')])
        writer.writerow(['File 2 (Client):', file_info2.get('original_filename')])
        writer.writerow([])
        
        # Statistics
        stats = comparison_result.get('statistics', {})
        writer.writerow(['Statistics'])
        for key, val in stats.items():
            writer.writerow([key.replace('_', ' ').title(), val])
        
        writer.writerow([])
        writer.writerow(['Overall Quality Score:', f"{comparison_result.get('quality_score', 0)}%"])
        
        return output.getvalue()
